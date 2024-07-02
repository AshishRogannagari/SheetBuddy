"""
SheetBuddy module for performing various EDA (Exploratory Data Analysis) operations on datasets.

"""
import time
from io import StringIO
from io import StringIO, BytesIO
import logging
import requests
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from scipy import stats
import numpy as np

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Expanded predefined dictionary for column descriptions
COLUMN_DESCRIPTIONS = {
    'name': 'The name of the individual or entity.',
    'age': 'The age of the individual.',
    'date': 'The date of the event or record.',
    'salary': 'The salary of the individual.',
    'country': 'The country associated with the record.',
    'city': 'The city associated with the record.',
    'gender': 'The gender of the individual.',
    'email': 'The email address of the individual or entity.',
    'phone': 'The phone number of the individual or entity.',
    'id': 'The unique identifier for the record.',
    'address': 'The street address of the individual or entity.',
    'zip code': 'The postal code for the address.',
    'state': 'The state or province associated with the record.',
    'department': 'The department within the organization.',
    'position': 'The job title or position of the individual.',
    'hire date': 'The date the individual was hired.',
    'birthdate': 'The birthdate of the individual.',
    'score': 'A score or rating associated with the individual or entity.',
    'status': 'The current status of the record.',
    'comment': 'Additional comments or notes about the record.',
    'category': 'The category or classification of the record.',
    'type': 'The type or nature of the record.',
    'quantity': 'The quantity associated with the record.',
    'price': 'The price or cost associated with the record.',
    'discount': 'Any discount applied to the price.',
    'total': 'The total amount for the record.',
    'description': 'A detailed description of the record.',
    'product': 'The product name or identifier.',
    'service': 'The service name or identifier.',
    'rating': 'A rating or review score for the record.',
    'review': 'A written review or feedback for the record.',
    'tag': 'A tag or label associated with the record.',
    'timestamp': 'The date and time when the record was created or updated.',
    'username': 'The username associated with the individual or entity.',
    'password': 'The password associated with the individual or entity.',
    'role': 'The role or permissions associated with the individual.',
    'group': 'The group or team associated with the individual.',
    'session': 'The session identifier for the record.',
}

class SheetBuddy:
    """
    A class to handle various EDA (Exploratory Data Analysis) operations on datasets.

    Attributes:
    -----------
    file_path_or_url : str
        The file path or URL to the dataset.
    """

    def __init__(self, file_path_or_url):
        """
        Initialize the SheetBuddy class with the file path or URL to the dataset.

        Parameters:
        -----------
        file_path_or_url : str
            The file path or URL to the dataset.
        """
        self.file_path_or_url = file_path_or_url
        self.data = self.read_data()

    def read_csv(self):
        """
        Read a CSV file into a DataFrame.

        Returns:
        --------
        DataFrame
            The dataset read from the CSV file.
        """
        try:
            if self.file_path_or_url.startswith("http"):
                response = requests.get(self.file_path_or_url, timeout=10)
                response.raise_for_status()
                data = StringIO(response.text)
                return pd.read_csv(data)
            return pd.read_csv(self.file_path_or_url)
        except (requests.RequestException, pd.errors.ParserError) as exc:
            logging.error("Error reading CSV file: %s", exc)
            return None

    def read_json(self):
        """
        Read JSON data from a URL or file into a DataFrame.

        Returns:
        --------
        DataFrame
            The dataset read from the JSON URL or file.
        """
        try:
            if self.file_path_or_url.startswith("http"):
                response = requests.get(self.file_path_or_url, timeout=10)
                response.raise_for_status()
                return pd.read_json(response.text)
            return pd.read_json(self.file_path_or_url)
        except (requests.RequestException, ValueError) as exc:
            logging.error("Error reading JSON file: %s", exc)
            return None

    def read_api(self):
        """
        Read data from an API into a DataFrame.

        Returns:
        --------
        DataFrame
            The dataset read from the API.
        """
        try:
            response = requests.get(self.file_path_or_url, timeout=10)
            response.raise_for_status()
            return pd.read_json(response.text)
        except (requests.RequestException, ValueError) as exc:
            logging.error("Error reading API data: %s", exc)
            return None

    def read_data(self):
        """
        Determine the file type and read the data accordingly.

        Returns:
        --------
        DataFrame
            The dataset read from the file or URL.
        """
        for _ in tqdm(range(100), desc="Loading data..."):
            time.sleep(0.01)  # Simulate some delay for progress bar
        if self.file_path_or_url.endswith('.csv'):
            return self.read_csv()
        if self.file_path_or_url.endswith('.json') or self.file_path_or_url.startswith("http"):
            return self.read_json()
        return self.read_api()

    def get_column_info(self):
        """
        Retrieve information about the columns in the dataset.

        Returns:
        --------
        DataFrame
            A DataFrame containing the column names, data types, descriptions,
            and whether they are categorical or numerical.
        """
        normalized_columns = [col.lower().replace('_', ' ') for col in self.data.columns]
        descriptions = [COLUMN_DESCRIPTIONS.get(col, None) for col in normalized_columns]
        return pd.DataFrame({
            'Column': self.data.columns,
            'Data Type': self.data.dtypes,
            'Description': descriptions,
            'Categorical/Numerical': [
                'Categorical' if self.data[col].dtype == 'object' else 'Numerical' for col in self.data.columns
            ]
        })

    def get_shape(self):
        """
        Retrieve the shape (dimensions) of the dataset.

        Returns:
        --------
        tuple
            The shape of the dataset as a tuple (rows, columns).
        """
        return self.data.shape

    def get_summary_statistics(self):
        """
        Retrieve summary statistics of the dataset.

        Returns:
        --------
        DataFrame
            The summary statistics of the dataset.
        """
        try:
            numeric_data = self.data.select_dtypes(include=['number'])
            if numeric_data.empty:
                return pd.DataFrame()
            return numeric_data.describe()
        except ValueError as exc:
            logging.error("Error generating summary statistics: %s", exc)
            return pd.DataFrame()

    def get_null_values(self):
        """
        Retrieve the count of null values for each column.

        Returns:
        --------
        Series
            A series containing the count of null values for each column.
        """
        return self.data.isnull().sum()

    def get_null_percentage(self):
        """
        Retrieve the percentage of null values for each column.

        Returns:
        --------
        Series
            A series containing the percentage of null values for each column.
        """
        return self.data.isnull().mean() * 100

    def get_standard_deviation(self):
        """
        Retrieve the standard deviation for numerical columns.

        Returns:
        --------
        Series
            A series containing the standard deviation for each numerical column.
        """
        try:
            numeric_data = self.data.select_dtypes(include=['number'])
            if numeric_data.empty:
                return pd.Series()
            return numeric_data.std()
        except ValueError as exc:
            logging.error("Error calculating standard deviation: %s", exc)
            return pd.Series()

    def get_unique_values(self):
        """
        Retrieve the count of unique values for each column.

        Returns:
        --------
        Series
            A series containing the count of unique values for each column.
        """
        return self.data.nunique()

    def get_most_frequent_values(self):
        """
        Retrieve the most frequent value for each column.

        Returns:
        --------
        Series
            A series containing the most frequent value for each column.
        """
        try:
            return self.data.mode().iloc[0]
        except ValueError as exc:
            logging.error("Error finding most frequent values: %s", exc)
            return pd.Series()

    def get_basic_math(self):
        """
        Retrieve basic mathematical calculations for numerical columns.

        Returns:
        --------
        DataFrame
            A DataFrame containing mean, median, mode, and range for each numerical column.
        """
        try:
            numerical_data = self.data.select_dtypes(include=['number'])
            if numerical_data.empty:
                return pd.DataFrame()
            return pd.DataFrame({
                'Mean': numerical_data.mean(),
                'Median': numerical_data.median(),
                'Mode': numerical_data.mode().iloc[0],
                'Range': numerical_data.max() - numerical_data.min()
            })
        except ValueError as exc:
            logging.error("Error calculating basic mathematics: %s", exc)
            return pd.DataFrame()

    def get_categorical_frequencies(self):
        """
        Retrieve the frequency of unique values for categorical columns.

        Returns:
        --------
        DataFrame
            A DataFrame containing the frequency of unique values for each categorical column.
        """
        try:
            categorical_data = self.data.select_dtypes(include=['object'])
            if categorical_data.empty:
                return pd.DataFrame()
            freq_dict = {col: categorical_data[col].value_counts() for col in categorical_data.columns}
            freq_df = pd.DataFrame(freq_dict)
            return freq_df
        except ValueError as exc:
            logging.error("Error calculating frequencies: %s", exc)
            return pd.DataFrame()

    def detect_outliers(self, method='z-score', threshold=3):
        """
        Detect outliers in the dataset.

        Parameters:
        -----------
        method : str
            The method to use for outlier detection ('z-score' or 'iqr').
        threshold : float
            The threshold value for detecting outliers.

        Returns:
        --------
        DataFrame
            A DataFrame indicating the presence of outliers in the dataset.
        """
        numeric_data = self.data.select_dtypes(include=['number'])
        outliers = pd.DataFrame(index=numeric_data.index, columns=numeric_data.columns)

        if method == 'z-score':
            z_scores = np.abs(stats.zscore(numeric_data))
            outliers = (z_scores > threshold)
        elif method == 'iqr':
            Q1 = numeric_data.quantile(0.25)
            Q3 = numeric_data.quantile(0.75)
            IQR = Q3 - Q1
            outliers = (numeric_data < (Q1 - 1.5 * IQR)) | (numeric_data > (Q3 + 1.5 * IQR))

        return outliers

    def style_excel_sheet(self, sheet):
        """
        Style an Excel sheet with header and row formatting.

        Parameters:
        -----------
        sheet : openpyxl.worksheet.worksheet.Worksheet
            The Excel sheet to style.
        """
        header_fill = PatternFill(start_color='4B4B4B', end_color='4B4B4B', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)  # White font color, bold, increased size
        center_alignment = Alignment(horizontal='center', vertical='center')
        row_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        border = Border(left=Side(style='thin', color='000000'), 
                        right=Side(style='thin', color='000000'), 
                        top=Side(style='thin', color='000000'), 
                        bottom=Side(style='thin', color='000000'))

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
                cell.border = border
                if cell.row == 1:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                elif cell.row % 2 == 0:  # Apply alternating row color
                    cell.fill = row_fill

    def add_text_heading(self, sheet, text, start_row):
        """
        Add a text heading to the Excel sheet.

        Parameters:
        -----------
        sheet : openpyxl.worksheet.worksheet.Worksheet
            The Excel sheet to add the text heading to.
        text : str
            The text for the heading.
        start_row : int
            The row to start placing the heading.
        """
        cell = sheet.cell(row=start_row, column=1)
        cell.value = text
        cell.font = Font(size=14, bold=True)
        return start_row + 1

    def add_histogram(self, sheet, data, column, start_row, color):
        """
        Add a histogram chart to the Excel sheet.

        Parameters:
        -----------
        sheet : openpyxl.worksheet.worksheet.Worksheet
            The Excel sheet to add the chart to.
        data : DataFrame
            The data for the histogram.
        column : str
            The column to plot in the histogram.
        start_row : int
            The row to start placing the chart.
        color : str
            The color to use for the histogram bars.
        """
        if data[column].dropna().empty:
            return start_row

        start_row = self.add_text_heading(sheet, f'Histogram of {column}', start_row)

        plt.figure(figsize=(10, 6))
        plt.hist(data[column].dropna(), bins=30, edgecolor='k', alpha=0.7, color=color)
        plt.title(f'Histogram of {column}')
        plt.xlabel(column)
        plt.ylabel('Frequency')

        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        plt.close()
        image_stream.seek(0)

        img = Image(image_stream)
        img.anchor = f'A{start_row}'
        sheet.add_image(img)

        return start_row + 25

    def add_correlation_heatmap(self, sheet, data, start_row):
        """
        Add a correlation heatmap to the Excel sheet.

        Parameters:
        -----------
        sheet : openpyxl.worksheet.worksheet.Worksheet
            The Excel sheet to add the heatmap to.
        data : DataFrame
            The data for the heatmap.
        start_row : int
            The row to start placing the heatmap.
        """
        numeric_data = data.select_dtypes(include=['number'])
        if numeric_data.empty:
            return start_row

        start_row = self.add_text_heading(sheet, 'Correlation Heatmap', start_row)

        plt.figure(figsize=(12, 10))
        correlation_matrix = numeric_data.corr()
        sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)
        plt.title('Correlation Heatmap')

        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        plt.close()
        image_stream.seek(0)

        img = Image(image_stream)
        img.anchor = f'A{start_row}'
        sheet.add_image(img)

        return start_row + 35

    def add_outliers_plot(self, sheet, data, column, start_row):
        """
        Add a boxplot to visualize outliers in the Excel sheet.

        Parameters:
        -----------
        sheet : openpyxl.worksheet.worksheet.Worksheet
            The Excel sheet to add the chart to.
        data : DataFrame
            The data for the boxplot.
        column : str
            The column to plot in the boxplot.
        start_row : int
            The row to start placing the chart.
        """
        if data[column].dropna().empty:
            return start_row

        start_row = self.add_text_heading(sheet, f'Outliers in {column}', start_row)

        plt.figure(figsize=(10, 6))
        sns.boxplot(data=data, x=column)
        plt.title(f'Outliers in {column}')
        plt.xlabel(column)

        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        plt.close()
        image_stream.seek(0)

        img = Image(image_stream)
        img.anchor = f'A{start_row}'
        sheet.add_image(img)

        return start_row + 25

    def add_dataset_info(self, workbook):
        """
        Add a Dataset Info sheet to the workbook.

        Parameters:
        -----------
        workbook : openpyxl.Workbook
            The workbook to add the Dataset Info sheet to.
        """
        sheet = workbook.create_sheet(title='Dataset Info', index=0)

        # Define the header and data
        headers = ['Name of the dataset', 'Format of the dataset', 'Number of rows', 'Number of columns', 'Dataset description', 'Data link']
        values = [
            'Dataset Name',  # Replace with actual dataset name if available
            'CSV' if self.file_path_or_url.endswith('.csv') else 'JSON',
            self.data.shape[0],
            self.data.shape[1],
            'Description of the dataset',  # Replace with actual description if available
            self.file_path_or_url if self.file_path_or_url.startswith('http') else 'N/A'
        ]

        # Add a title
        sheet['A1'] = 'Data Summary'
        sheet['A1'].font = Font(size=18, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:B1')

        # Set column widths
        column_widths = [30, 45]

        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width

        # Add headers and values
        for row, (header, value) in enumerate(zip(headers, values), start=2):
            cell_header = sheet.cell(row=row, column=1, value=header)
            cell_header.font = Font(bold=True, size=12)
            cell_header.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            cell_header.alignment = Alignment(horizontal='left', vertical='center')

            cell_value = sheet.cell(row=row, column=2, value=value)
            cell_value.font = Font(size=12)
            cell_value.alignment = Alignment(horizontal='left', vertical='center')
            if headers[row-2] == 'Data link' and value != 'N/A':
                cell_value.hyperlink = value
                cell_value.style = "Hyperlink"

        # Apply black border to the entire table
        border = Border(left=Side(style='thin', color='000000'), 
                        right=Side(style='thin', color='000000'), 
                        top=Side(style='thin', color='000000'), 
                        bottom=Side(style='thin', color='000000'))

        for row in sheet.iter_rows(min_row=1, max_row=len(headers) + 2, min_col=1, max_col=2):
            for cell in row:
                cell.border = border

    def generate_report(self, file_name):
        """
        Generate an EDA report and save it to an Excel file.

        Parameters:
        -----------
        file_name : str
            The name of the output Excel file.
        """
        try:
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                # Add Dataset Info sheet
                workbook = writer.book
                self.add_dataset_info(workbook)

                # Add Column Info sheet
                column_info = self.get_column_info()
                column_info.to_excel(writer, sheet_name='Column Info', index=False)

                # Add Shape Info sheet
                shape_info = pd.DataFrame({'Shape': [self.get_shape()]})
                shape_info.to_excel(writer, sheet_name='Shape Info', index=False)

                # Add Summary Statistics sheet
                summary_stats = self.get_summary_statistics().reset_index()
                summary_stats.to_excel(writer, sheet_name='Summary Statistics', index=False)

                # Add Null Values sheet
                null_values = self.get_null_values().reset_index()
                null_values.columns = ['Column', 'Null Values']
                null_values.to_excel(writer, sheet_name='Null Values', index=False)

                # Add Null Percentage sheet
                null_percentage = self.get_null_percentage().reset_index()
                null_percentage.columns = ['Column', 'Null Percentage']
                null_percentage.to_excel(writer, sheet_name='Null Percentage', index=False)

                # Add Standard Deviation sheet
                std_dev = self.get_standard_deviation().reset_index()
                std_dev.columns = ['Column', 'Standard Deviation']
                std_dev.to_excel(writer, sheet_name='Standard Deviation', index=False)

                # Add Unique Values sheet
                unique_values = self.get_unique_values().reset_index()
                unique_values.columns = ['Column', 'Unique Values']
                unique_values.to_excel(writer, sheet_name='Unique Values', index=False)

                # Add Most Frequent Values sheet
                most_frequent_values = self.get_most_frequent_values().reset_index()
                most_frequent_values.columns = ['Column', 'Most Frequent Value']
                most_frequent_values.to_excel(writer, sheet_name='Most Frequent Values', index=False)

                # Add Categorical Frequencies sheet
                categorical_frequencies = self.get_categorical_frequencies().reset_index()
                categorical_frequencies.to_excel(writer, sheet_name='Categorical Frequencies', index=False)

                # Add Correlation Matrix sheet
                numerical_data = self.data.select_dtypes(include=['number'])
                if not numerical_data.empty:
                    correlation_matrix = numerical_data.corr()
                    correlation_matrix.to_excel(writer, sheet_name='Correlation Matrix')

                # Add Basic Mathematics sheet
                basic_math = self.get_basic_math().reset_index()
                if not basic_math.empty:
                    basic_math.to_excel(writer, sheet_name='Basic Mathematics', index=False)

                # Style the sheets
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    self.style_excel_sheet(sheet)

                # Add EDA sheet with visualizations
                eda_sheet = workbook.create_sheet(title='Visualizations')
                start_row = 1
                colors = ['blue', 'green', 'red', 'purple', 'orange']
                for i, column in enumerate(numerical_data.columns):
                    color = colors[i % len(colors)]
                    start_row = self.add_histogram(eda_sheet, self.data, column, start_row, color)
                    start_row += 6  # Additional spacing after each chart

                for column in numerical_data.columns:
                    start_row = self.add_outliers_plot(eda_sheet, self.data, column, start_row)
                    start_row += 6  # Additional spacing after each chart

                self.add_correlation_heatmap(eda_sheet, self.data, start_row)

        except Exception as e:
            logging.error("An error occurred during report generation: %s", e)

    def generate_eda_report(self, output_file_name):
        """
        Load data and generate the EDA report.

        Parameters:
        -----------
        output_file_name : str
            The name of the output Excel file.
        """
        logging.info("Loading data...")
        if self.data is not None:
            logging.info("Data loaded successfully.")
            logging.info("Generating report...")
            for _ in tqdm(range(100), desc="Generating report..."):
                time.sleep(0.01)  # Simulate some delay for progress bar
            self.generate_report(output_file_name)
            logging.info("Report generated: %s", output_file_name)
        else:
            logging.error("Failed to load data. Report generation aborted.")
