import time
import pandas as pd
import requests
from io import StringIO
from tqdm import tqdm
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Expanded predefined dictionary for column descriptions
COLUMN_DESCRIPTIONS = {
    'name': 'The name of the individual or entity.',
    'age': 'The age of the individual.',
    'date': 'The date of the event or record.',
    'salary': 'The salary of the individual.',
    'country': 'The country associated with the record.',
    'city': 'The city associated with the record.',
    'gender': 'The gender of the individual.',
    'email': 'The email address of the individual or entity.',
    'phone': 'The phone number of the individual or entity.',
    'id': 'The unique identifier for the record.',
    'address': 'The street address of the individual or entity.',
    'zip code': 'The postal code for the address.',
    'state': 'The state or province associated with the record.',
    'department': 'The department within the organization.',
    'position': 'The job title or position of the individual.',
    'hire date': 'The date the individual was hired.',
    'birthdate': 'The birthdate of the individual.',
    'score': 'A score or rating associated with the individual or entity.',
    'status': 'The current status of the record.',
    'comment': 'Additional comments or notes about the record.',
    'category': 'The category or classification of the record.',
    'type': 'The type or nature of the record.',
    'quantity': 'The quantity associated with the record.',
    'price': 'The price or cost associated with the record.',
    'discount': 'Any discount applied to the price.',
    'total': 'The total amount for the record.',
    'description': 'A detailed description of the record.',
    'product': 'The product name or identifier.',
    'service': 'The service name or identifier.',
    'rating': 'A rating or review score for the record.',
    'review': 'A written review or feedback for the record.',
    'tag': 'A tag or label associated with the record.',
    'timestamp': 'The date and time when the record was created or updated.',
    'username': 'The username associated with the individual or entity.',
    'password': 'The password associated with the individual or entity.',
    'role': 'The role or permissions associated with the individual.',
    'group': 'The group or team associated with the individual.',
    'session': 'The session identifier for the record.',
}

class SheetBuddy:
    def __init__(self, file_path_or_url):
        self.file_path_or_url = file_path_or_url
        self.data = self.read_data()

    def read_csv(self):
        try:
            if self.file_path_or_url.startswith("http"):
                response = requests.get(self.file_path_or_url)
                response.raise_for_status()
                data = StringIO(response.text)
                return pd.read_csv(data)
            else:
                return pd.read_csv(self.file_path_or_url)
        except Exception as e:
            logging.error(f"Error reading CSV file: {e}")
            return None

    def read_json(self):
        try:
            if self.file_path_or_url.startswith("http"):
                response = requests.get(self.file_path_or_url)
                response.raise_for_status()
                return pd.read_json(response.text)
            else:
                return pd.read_json(self.file_path_or_url)
        except Exception as e:
            logging.error(f"Error reading JSON file: {e}")
            return None

    def read_api(self):
        try:
            response = requests.get(self.file_path_or_url)
            response.raise_for_status()
            return pd.read_json(response.text)
        except Exception as e:
            logging.error(f"Error reading API data: {e}")
            return None

    def read_data(self):
        for _ in tqdm(range(100), desc="Loading data..."):
            time.sleep(0.01)  # Simulate some delay for progress bar
        if self.file_path_or_url.endswith('.csv'):
            return self.read_csv()
        elif self.file_path_or_url.endswith('.json') or self.file_path_or_url.startswith("http"):
            return self.read_json()
        else:
            return self.read_api()

    def get_column_info(self):
        normalized_columns = [col.lower().replace('_', ' ') for col in self.data.columns]
        descriptions = [COLUMN_DESCRIPTIONS.get(col, None) for col in normalized_columns]
        return pd.DataFrame({
            'Column': self.data.columns,
            'Data Type': self.data.dtypes,
            'Description': descriptions,
            'Categorical/Numerical': ['Categorical' if self.data[col].dtype == 'object' else 'Numerical' for col in self.data.columns]
        })

    def get_shape(self):
        return self.data.shape

    def get_summary_statistics(self):
        try:
            return self.data.describe()
        except Exception as e:
            logging.error(f"Error generating summary statistics: {e}")
            return pd.DataFrame()

    def get_null_values(self):
        return self.data.isnull().sum()

    def get_null_percentage(self):
        return self.data.isnull().mean() * 100

    def get_standard_deviation(self):
        try:
            return self.data.std()
        except Exception as e:
            logging.error(f"Error calculating standard deviation: {e}")
            return pd.Series()

    def get_unique_values(self):
        return self.data.nunique()

    def get_most_frequent_values(self):
        try:
            return self.data.mode().iloc[0]
        except Exception as e:
            logging.error(f"Error finding most frequent values: {e}")
            return pd.Series()

    def apply_conditional_formatting(self, sheet, column, fill_color):
        for cell in sheet[column]:
            if cell.row == 1:  # Skip header row
                continue
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    def style_excel_sheet(self, sheet):
        header_fill = PatternFill(start_color='4B4B4B', end_color='4B4B4B', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)  # White font color, bold, increased size
        center_alignment = Alignment(horizontal='center', vertical='center')
        row_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
                if cell.row == 1:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                elif cell.row % 2 == 0:  # Apply alternating row color
                    cell.fill = row_fill

    def generate_report(self, file_name):
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            column_info = self.get_column_info()
            column_info.to_excel(writer, sheet_name='Column Info', index=False)

            shape_info = pd.DataFrame({'Shape': [self.get_shape()]})
            shape_info.to_excel(writer, sheet_name='Shape Info', index=False)

            numeric_df = self.data.select_dtypes(include=['number'])
            summary_stats = self.get_summary_statistics().reset_index()
            summary_stats.to_excel(writer, sheet_name='Summary Statistics', index=False)

            null_values = self.get_null_values().reset_index()
            null_values.columns = ['Column', 'Null Values']
            null_values.to_excel(writer, sheet_name='Null Values', index=False)

            null_percentage = self.get_null_percentage().reset_index()
            null_percentage.columns = ['Column', 'Null Percentage']
            null_percentage.to_excel(writer, sheet_name='Null Percentage', index=False)

            std_dev = self.get_standard_deviation().reset_index()
            std_dev.columns = ['Column', 'Standard Deviation']
            std_dev.to_excel(writer, sheet_name='Standard Deviation', index=False)

            unique_values = self.get_unique_values().reset_index()
            unique_values.columns = ['Column', 'Unique Values']
            unique_values.to_excel(writer, sheet_name='Unique Values', index=False)

            most_frequent_values = self.get_most_frequent_values().reset_index()
            most_frequent_values.columns = ['Column', 'Most Frequent Value']
            most_frequent_values.to_excel(writer, sheet_name='Most Frequent Values', index=False)

            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self.style_excel_sheet(sheet)

    def generate_eda_report(self, output_file_name):
        logging.info("Loading data...")
        if self.data is not None:
            logging.info("Data loaded successfully.")
            logging.info("Generating report...")
            for _ in tqdm(range(100), desc="Generating report..."):
                time.sleep(0.01)  # Simulate some delay for progress bar
            self.generate_report(output_file_name)
            logging.info(f"Report generated: {output_file_name}")
        else:
            logging.error("Failed to load data. Report generation aborted.")

